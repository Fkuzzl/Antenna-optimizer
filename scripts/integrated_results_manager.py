#!/usr/bin/env python3
"""
Integrated Results Manager for MATLAB-HFSS Optimization
Manages CSV to Excel integration for S11, AR, and Gain data across iterations.
"""

import os
import sys
import pandas as pd
import argparse
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class IntegratedResultsManager:
    def __init__(self, project_path, excel_path=None):
        """Initialize the Integrated Results Manager."""
        self.project_path = Path(project_path)
        self.excel_path = Path(excel_path) if excel_path else self.project_path / "Integrated_Results.xlsx"
        self.optimization_data_path = self.project_path / "Optimization" / "data"
        
        # Data type patterns for file detection
        self.data_patterns = {
            'S11': {'pattern': r'S11_(\d+)\.csv', 'sheet': 'S11_Data', 'value_col': 'S11_dB'},
            'AR': {'pattern': r'AR_(\d+)\.csv', 'sheet': 'AR_Data', 'value_col': 'AR'},
            'Gain': {'pattern': r'Gain_(\d+)\.csv', 'sheet': 'Gain_Data', 'value_col': 'Gain_dBi'}
        }

    def scan_csv_files(self):
        """Scan for CSV files in the optimization data directory."""
        import re
        
        if not self.optimization_data_path.exists():
            print(f"   Optimization data path not found: {self.optimization_data_path}")
            return {}
        
        csv_files = {}
        
        for data_type, config in self.data_patterns.items():
            csv_files[data_type] = []
            pattern = re.compile(config['pattern'])
            
            for file_path in self.optimization_data_path.glob("*.csv"):
                match = pattern.match(file_path.name)
                if match:
                    iteration = int(match.group(1))
                    csv_files[data_type].append({
                        'filepath': file_path,
                        'filename': file_path.name,
                        'iteration': iteration
                    })
            
            # Sort by iteration number
            csv_files[data_type].sort(key=lambda x: x['iteration'])
        
        return csv_files

    def read_csv_file(self, filepath):
        """Read CSV file and standardize column names."""
        try:
            df = pd.read_csv(filepath)
            
            # Detect and standardize column names
            freq_col = None
            value_col = None
            
            for col in df.columns:
                col_lower = col.lower()
                if 'freq' in col_lower:
                    freq_col = col
                elif any(keyword in col_lower for keyword in ['s(1,1)', 's11', 'ar', 'gain', 'db(']):
                    value_col = col
            
            if freq_col is None or value_col is None:
                print(f"[WARNING] Could not identify columns in {filepath}")
                return None
            
            # Standardize column names
            df = df.rename(columns={freq_col: 'Frequency_GHz', value_col: 'Value'})
            return df[['Frequency_GHz', 'Value']]
            
        except Exception as e:
            print(f"[ERROR] Error reading {filepath}: {str(e)}")
            return None

    def create_integrated_excel(self):
        """Create integrated Excel file from all CSV files."""
        try:
            csv_files = self.scan_csv_files()
            
            # Create workbook regardless of CSV files availability
            if self.excel_path.exists():
                print(f"   Loading existing Excel file: {self.excel_path}")
                workbook = load_workbook(self.excel_path)
                # Clear existing sheets
                for sheet_name in list(workbook.sheetnames):
                    del workbook[sheet_name]
            else:
                print(f"   Creating new Excel file: {self.excel_path}")
                workbook = Workbook()
                # Remove default sheet
                workbook.remove(workbook.active)
            
            if not any(csv_files.values()):
                print("   No CSV files found - creating empty Excel with basic structure")
                # Create empty sheets with headers for each data type
                for data_type, config in self.data_patterns.items():
                    sheet_name = config['sheet']
                    value_col_name = config['value_col']
                    
                    # Create worksheet with headers
                    worksheet = workbook.create_sheet(title=sheet_name)
                    headers = ['Iteration', 'Frequency_GHz', value_col_name]
                    worksheet.append(headers)
                    print(f"   Created empty sheet '{sheet_name}' with headers")
                
                # Save empty workbook
                workbook.save(self.excel_path)
                print(f"   Empty integrated Excel file created: {self.excel_path}")
                return True
            
            # Process each data type (workbook already created above)
            for data_type, files in csv_files.items():
                if not files:
                    continue
                
                sheet_name = self.data_patterns[data_type]['sheet']
                value_col_name = self.data_patterns[data_type]['value_col']
                
                # Combine all iterations for this data type
                combined_data = []
                
                for file_info in files:
                    df = self.read_csv_file(file_info['filepath'])
                    if df is not None:
                        df['Iteration'] = file_info['iteration']
                        df = df.rename(columns={'Value': value_col_name})
                        df = df[['Iteration', 'Frequency_GHz', value_col_name]]
                        combined_data.append(df)
                        print(f"[OK] Processed {file_info['filename']}: {len(df)} rows")
                
                if combined_data:
                    combined_df = pd.concat(combined_data, ignore_index=True)
                    combined_df = combined_df.sort_values(['Iteration', 'Frequency_GHz'])
                    
                    # Create worksheet
                    worksheet = workbook.create_sheet(title=sheet_name)
                    
                    # Write data to worksheet
                    for row in dataframe_to_rows(combined_df, index=False, header=True):
                        worksheet.append(row)
                    
                    print(f"   Created sheet '{sheet_name}' with {len(combined_df)} total rows")
            
            # Save workbook
            workbook.save(self.excel_path)
            print(f"   Integrated Excel file saved: {self.excel_path}")
            return True
            
        except Exception as e:
            print(f"[ERROR] Error creating integrated Excel: {str(e)}")
            return False

    def update_integrated_excel(self, iteration=None):
        """Update integrated Excel file with new data."""
        # For now, just recreate the entire file
        return self.create_integrated_excel()

    def clear_integrated_excel(self):
        """Clear/delete the integrated Excel file."""
        try:
            if self.excel_path.exists():
                self.excel_path.unlink()
                print(f"   Cleared integrated Excel file: {self.excel_path}")
                return True
            else:
                print(f"   Excel file does not exist: {self.excel_path}")
                return False
        except Exception as e:
            print(f"[ERROR] Error clearing Excel file: {str(e)}")
            return False

    def get_summary(self):
        """Get summary information about the integrated Excel file."""
        try:
            if not self.excel_path.exists():
                return {"exists": False, "path": str(self.excel_path)}
            
            workbook = load_workbook(self.excel_path, data_only=True)
            sheets_info = {}
            max_iteration = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = worksheet.max_row
                cols = worksheet.max_column
                sheets_info[sheet_name] = {"rows": rows - 1, "columns": cols}  # -1 for header
                
                # Try to find max iteration
                try:
                    for row in worksheet.iter_rows(min_row=2, max_row=min(rows, 100), min_col=1, max_col=1):
                        cell_value = row[0].value
                        if cell_value and isinstance(cell_value, (int, float)):
                            max_iteration = max(max_iteration, int(cell_value))
                except:
                    pass
            
            return {
                "exists": True,
                "path": str(self.excel_path),
                "sheets": sheets_info,
                "total_iterations": max_iteration
            }
            
        except Exception as e:
            return {"exists": False, "error": str(e), "path": str(self.excel_path)}

def main():
    parser = argparse.ArgumentParser(description='Integrated Results Manager for MATLAB-HFSS Optimization')
    parser.add_argument('action', choices=['create', 'update', 'clear', 'summary'],
                       help='Action to perform')
    parser.add_argument('--project-path', required=True,
                       help='Path to MATLAB project directory')
    parser.add_argument('--excel-path', 
                       help='Custom path for integrated Excel file')
    parser.add_argument('--iteration', type=int,
                       help='Specific iteration to process (for update action)')
    
    args = parser.parse_args()
    
    # Initialize manager
    manager = IntegratedResultsManager(args.project_path, args.excel_path)
    
    # Only print header for non-summary actions (summary needs clean JSON output)
    if args.action != 'summary':
        print(f">> Integrated Results Manager")
        print(f"   Project path: {args.project_path}")
        print(f"   Excel path: {manager.excel_path}")
        print(f"   Action: {args.action}")
    
    if args.action == 'create':
        success = manager.create_integrated_excel()
        if success:
            print("[OK] Integrated Excel file created successfully")
        else:
            print("[ERROR] Failed to create integrated Excel file")
    
    elif args.action == 'update':
        success = manager.update_integrated_excel(args.iteration)
        if success:
            print(f"[OK] Excel file updated successfully" + 
                  (f" (iteration {args.iteration})" if args.iteration else ""))
        else:
            print("[ERROR] Failed to update Excel file")
    
    elif args.action == 'clear':
        success = manager.clear_integrated_excel()
        if success:
            print("[OK] Excel file cleared successfully")
        else:
            print("[ERROR] Failed to clear Excel file")
    
    elif args.action == 'summary':
        summary = manager.get_summary()
        # For summary action, output only JSON (no extra text for server parsing)
        print(json.dumps(summary, indent=2))

if __name__ == '__main__':
    main()