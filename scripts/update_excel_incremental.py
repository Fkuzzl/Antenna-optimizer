#!/usr/bin/env python3
"""
Simple incremental Excel updater - handles multiple missing iterations
Checks Excel file, finds missing iterations from CSV files, and appends them.
"""

import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
import re
import argparse

def get_last_iteration_in_excel(excel_path):
    """Get the highest iteration number currently in Excel."""
    try:
        wb = load_workbook(excel_path, read_only=True)
        max_iteration = 0
        
        # Check all sheets for the highest iteration number
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] is not None and isinstance(row[0], (int, float)):
                    max_iteration = max(max_iteration, int(row[0]))
        
        wb.close()
        return max_iteration
    except Exception as e:
        print(f"[ERROR] Could not read Excel file: {e}")
        return 0

def find_csv_iterations(data_path, start_iteration):
    """Find all CSV iterations greater than start_iteration."""
    data_patterns = {
        'S11': r'S11_(\d+)\.csv',
        'AR': r'AR_(\d+)\.csv',
        'Gain': r'Gain_(\d+)\.csv'
    }
    
    iterations = set()
    
    for data_type, pattern in data_patterns.items():
        regex = re.compile(pattern)
        for file in Path(data_path).glob("*.csv"):
            match = regex.match(file.name)
            if match:
                iteration = int(match.group(1))
                if iteration > start_iteration:
                    iterations.add(iteration)
    
    return sorted(list(iterations))

def read_csv_standardized(filepath):
    """Read CSV and return standardized DataFrame."""
    try:
        df = pd.read_csv(filepath)
        
        # Find frequency and value columns
        freq_col = None
        value_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'freq' in col_lower:
                freq_col = col
            elif any(kw in col_lower for kw in ['s(1,1)', 's11', 'ar', 'gain', 'db(']):
                value_col = col
        
        if freq_col and value_col:
            return df[[freq_col, value_col]].rename(columns={
                freq_col: 'Frequency_GHz',
                value_col: 'Value'
            })
    except Exception as e:
        print(f"[ERROR] Reading {filepath}: {e}")
    
    return None

def append_iteration_to_excel(wb, data_path, iteration):
    """Append a single iteration's data to an already-opened workbook."""
    data_configs = {
        'S11': {'pattern': f'S11_{iteration}.csv', 'sheet': 'S11_Data', 'col': 'S11_dB'},
        'AR': {'pattern': f'AR_{iteration}.csv', 'sheet': 'AR_Data', 'col': 'AR'},
        'Gain': {'pattern': f'Gain_{iteration}.csv', 'sheet': 'Gain_Data', 'col': 'Gain_dBi'}
    }
    
    for data_type, config in data_configs.items():
        csv_file = Path(data_path) / config['pattern']
        
        if not csv_file.exists():
            continue
        
        df = read_csv_standardized(csv_file)
        if df is None:
            continue
        
        # Add iteration column
        df['Iteration'] = iteration
        df = df.rename(columns={'Value': config['col']})
        df = df[['Iteration', 'Frequency_GHz', config['col']]]
        
        # Get or create sheet
        if config['sheet'] not in wb.sheetnames:
            ws = wb.create_sheet(config['sheet'])
            ws.append(['Iteration', 'Frequency_GHz', config['col']])
        else:
            ws = wb[config['sheet']]
        
        # Append rows
        for _, row in df.iterrows():
            ws.append([row['Iteration'], row['Frequency_GHz'], row[config['col']]])

def update_excel_incremental(project_path):
    """Main update function - finds and appends all missing iterations."""
    project_path = Path(project_path)
    excel_path = project_path / "Integrated_Results.xlsx"
    data_path = project_path / "Optimization" / "data"
    
    if not data_path.exists():
        print(f"[ERROR] Data path not found: {data_path}")
        return False
    
    # Check if Excel file exists, create if not
    if not excel_path.exists():
        # Silently create the Excel file
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create empty sheets with headers
            data_configs = {
                'S11_Data': 'S11_dB',
                'AR_Data': 'AR',
                'Gain_Data': 'Gain_dBi'
            }
            
            for sheet_name, value_col in data_configs.items():
                ws = wb.create_sheet(sheet_name)
                ws.append(['Iteration', 'Frequency_GHz', value_col])
            
            wb.save(excel_path)
            wb.close()
            print(f"Created Excel file with structure")
        except Exception as e:
            print(f"[ERROR] Failed to create Excel file: {e}")
            return False
    
    # Get current state
    last_excel_iter = get_last_iteration_in_excel(excel_path)
    
    # Find missing iterations
    missing_iterations = find_csv_iterations(data_path, last_excel_iter)
    
    if not missing_iterations:
        print("Results up to date")
        return True
    
    print(f"Processing {len(missing_iterations)} iterations")
    
    # Open Excel once, add all iterations, save once
    wb = None
    try:
        wb = load_workbook(excel_path)
        
        # Append each iteration to the open workbook
        for i, iteration in enumerate(missing_iterations, 1):
            try:
                append_iteration_to_excel(wb, data_path, iteration)
            except Exception as e:
                print(f"[ERROR] {e}")
                return False
        
        # Save once at the end
        wb.save(excel_path)
        
    except Exception as e:
        print(f"[ERROR] Failed to update Excel: {e}")
        return False
    finally:
        if wb:
            wb.close()
    
    print(f"Results loaded successfully - {missing_iterations[-1]} total iterations")
    return True

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Update Excel with missing iterations")
    parser.add_argument("--project-path", required=True, help="Path to project folder")
    args = parser.parse_args()
    
    success = update_excel_incremental(args.project_path)
    sys.exit(0 if success else 1)
